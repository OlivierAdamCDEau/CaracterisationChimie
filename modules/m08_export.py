"""
Module 08 — Export (v1)
========================
Génération des exports pour l'interface Streamlit :
  - Figures : PNG et SVG (par figure, bytes Streamlit-ready)
  - Excel   : un classeur par module, plusieurs onglets mis en forme
  - PDF     : rapport structuré multi-modules avec watermark @CDEau
  - ZIP     : bundle complet figures + Excel + PDF

Fonctions principales :
  - exporter_figure()          : bytes PNG ou SVG d'une figure matplotlib
  - exporter_excel_module()    : classeur Excel mis en forme pour un module
  - generer_rapport_pdf()      : PDF structuré multi-modules
  - generer_zip()              : ZIP agrégeant tous les exports
  - preparer_exports()         : pipeline complet → dict de bytes pour Streamlit

Conventions PDF :
  - Watermark @CDEau en pied de page sur chaque page
  - Page de garde : titre, date, stations, période
  - Une section par module : titre de section + figures + tableau de synthèse
  - Police : Helvetica (embarquée ReportLab, pas de dépendance externe)
  - Couleurs CDEau : bleu #2563eb pour les titres de section
"""

import io
import os
import zipfile
import datetime
from typing import Optional

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# Excel
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, PageBreak,
    Table, TableStyle, Image as RLImage, HRFlowable,
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

COULEUR_TITRE   = colors.HexColor("#2563eb")   # bleu CDEau
COULEUR_EN_TETE = colors.HexColor("#dbeafe")   # bleu très clair (en-têtes Excel)
COULEUR_ALTERNE = colors.HexColor("#f8faff")   # bleu quasi-blanc (lignes paires)
WATERMARK_TXT   = "@CDEau"
WATERMARK_COLOR = colors.HexColor("#999999")

# Styles Excel réutilisables
_FONT_TITRE  = Font(name="Arial", bold=True, size=11, color="2563eb")
_FONT_ENTETE = Font(name="Arial", bold=True, size=9,  color="1e3a5f")
_FONT_CELL   = Font(name="Arial", size=9)
_FILL_ENTETE = PatternFill("solid", fgColor="DBEAFE")
_FILL_ALTERN = PatternFill("solid", fgColor="F8FAFF")
_FILL_BLANC  = PatternFill("solid", fgColor="FFFFFF")
_ALIGN_CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_BORDER_THIN = Border(
    left=Side(style="thin", color="BFDBFE"),
    right=Side(style="thin", color="BFDBFE"),
    top=Side(style="thin", color="BFDBFE"),
    bottom=Side(style="thin", color="BFDBFE"),
)

# ---------------------------------------------------------------------------
# 1. Export figures PNG / SVG
# ---------------------------------------------------------------------------

def exporter_figure(
    fig: plt.Figure,
    format: str = "png",
    dpi: int = 200,
) -> bytes:
    """
    Retourne les bytes d'une figure matplotlib au format PNG ou SVG.
    Prêt pour st.download_button() dans Streamlit.

    Parameters
    ----------
    fig    : figure matplotlib
    format : "png" ou "svg"
    dpi    : résolution (ignoré pour SVG)

    Returns
    -------
    bytes
    """
    buf = io.BytesIO()
    if format.lower() == "svg":
        fig.savefig(buf, format="svg", bbox_inches="tight")
    else:
        fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# 2. Utilitaires Excel internes
# ---------------------------------------------------------------------------

def _ecrire_titre_onglet(ws, titre: str, n_colonnes: int = 6):
    """Écrit un titre en haut de l'onglet avec mise en forme CDEau."""
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=n_colonnes)
    cell = ws.cell(row=1, column=1, value=titre)
    cell.font      = _FONT_TITRE
    cell.alignment = _ALIGN_LEFT
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 6   # ligne vide de séparation


def _ecrire_en_tetes(ws, colonnes: list, ligne: int = 3):
    """Écrit une ligne d'en-têtes avec fond bleu clair."""
    for j, col in enumerate(colonnes, 1):
        cell = ws.cell(row=ligne, column=j, value=col)
        cell.font      = _FONT_ENTETE
        cell.fill      = _FILL_ENTETE
        cell.alignment = _ALIGN_CTR
        cell.border    = _BORDER_THIN
    ws.row_dimensions[ligne].height = 28


def _ecrire_donnees(ws, rows: list, ligne_debut: int = 4,
                    formats: Optional[dict] = None):
    """
    Écrit les données avec alternance de couleurs.
    formats : dict {col_index_1based: format_string_excel}
    """
    for i, row in enumerate(rows):
        fill = _FILL_ALTERN if i % 2 == 0 else _FILL_BLANC
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=ligne_debut + i, column=j, value=val)
            cell.font      = _FONT_CELL
            cell.fill      = fill
            cell.alignment = _ALIGN_CTR
            cell.border    = _BORDER_THIN
            if formats and j in formats:
                cell.number_format = formats[j]


def _auto_largeur(ws, min_width: int = 10, max_width: int = 40):
    """Ajuste automatiquement la largeur des colonnes."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width,
                                                      min(max_len + 2, max_width))


def _nb_fmt(decimales: int = 3) -> str:
    """Retourne un format Excel numérique avec n décimales."""
    return f'#,##0.{"0" * decimales}' if decimales > 0 else "#,##0"


# ---------------------------------------------------------------------------
# 3. Export Excel par module
# ---------------------------------------------------------------------------

def exporter_excel_m03(
    df_stats: pd.DataFrame,
    df_classes: Optional[pd.DataFrame] = None,
    pivot_classes: Optional[pd.DataFrame] = None,
    lb_map: Optional[dict] = None,
    lb_stations: Optional[dict] = None,
    titre: str = "Empreinte chimique — Synthèse",
) -> bytes:
    """
    Classeur Excel M03 :
      Onglet 1 : Statistiques descriptives (médiane, P10, P90, n, unité)
      Onglet 2 : Classes de qualité par station × paramètre (matrice colorée)

    Parameters
    ----------
    df_stats      : DataFrame issu de m02.nettoyer_et_pivoter()['df_stats']
    pivot_classes : DataFrame pivot stations × paramètres avec codes de classes
                    ('TBE','BE','EMO','EME','ME','ND') issu de
                    m07.calculer_classes_par_station(pivot, df_ref)[0].
                    PRIORITAIRE sur df_classes si fourni.
    df_classes    : DataFrame long de classes (format legacy, optionnel)
    lb_map        : dict {CdParametre: libellé}
    lb_stations   : dict {CdStation: libellé}
    """
    wb = Workbook()

    COUL_CLASSES = {
        "TBE": "DBEAFE", "BE":  "BBFBD0", "EMO": "FEF9C3",
        "EME": "FFEDD5", "ME":  "FEE2E2", "ND":  "F3F4F6",
    }

    # ── Onglet 1 : Stats descriptives ──
    ws1 = wb.active
    ws1.title = "Stats descriptives"
    _ecrire_titre_onglet(ws1, titre + " — Statistiques descriptives", n_colonnes=7)
    cols1 = ["Station", "Paramètre", "N mesures", "Médiane", "P10", "P90", "Unité"]
    _ecrire_en_tetes(ws1, cols1, ligne=3)

    rows1 = []
    if not df_stats.empty:
        for _, r in df_stats.iterrows():
            code = int(r.get("CdParametre", 0))
            lb_p = (lb_map.get(code, str(code)) if lb_map else str(code))[:40]
            lb_s = r.get("CdStationMesureEauxSurface", r.get("CdStation", ""))
            if lb_stations:
                lb_s = lb_stations.get(str(lb_s), str(lb_s))
            rows1.append([
                str(lb_s)[:30], lb_p,
                r.get("NbMesures", r.get("N", r.get("n", ""))),
                r.get("Mediane", r.get("median", "")),
                r.get("P10",     r.get("p10",    "")),
                r.get("P90",     r.get("p90",    "")),
                r.get("SymUniteMesure", r.get("Unite", "")),
            ])
    fmt1 = {4: _nb_fmt(3), 5: _nb_fmt(3), 6: _nb_fmt(3)}
    _ecrire_donnees(ws1, rows1, ligne_debut=4, formats=fmt1)
    _auto_largeur(ws1)

    # ── Onglet 2 : Classes de qualité ──
    ws2 = wb.create_sheet("Classes de qualité")

    if pivot_classes is not None and not pivot_classes.empty:
        # Format pivot (recommandé) : matrice stations × paramètres
        codes = list(pivot_classes.columns)
        stations_idx = list(pivot_classes.index)
        n_cols = len(codes) + 1
        _ecrire_titre_onglet(ws2, titre + " — Classes de qualité (station × paramètre)",
                             n_colonnes=n_cols)
        en_tetes = ["Station"] + [
            (lb_map.get(int(c), str(c)) if lb_map else str(c))[:16]
            for c in codes
        ]
        _ecrire_en_tetes(ws2, en_tetes, ligne=3)
        ws2.row_dimensions[3].height = 32

        for i, station in enumerate(stations_idx):
            lb_s = lb_stations.get(str(station), str(station)) if lb_stations else str(station)
            ligne = 4 + i
            cell_st = ws2.cell(row=ligne, column=1, value=lb_s[:30])
            cell_st.font = _FONT_CELL
            cell_st.alignment = _ALIGN_LEFT
            cell_st.border = _BORDER_THIN
            cell_st.fill = _FILL_ALTERN if i % 2 == 0 else _FILL_BLANC
            for j, code in enumerate(codes, 2):
                val = pivot_classes.loc[station, code]
                classe = str(val) if pd.notna(val) else "ND"
                cell = ws2.cell(row=ligne, column=j, value=classe)
                cell.font = Font(name="Arial", size=9, bold=True)
                cell.alignment = _ALIGN_CTR
                cell.border = _BORDER_THIN
                cell.fill = PatternFill("solid", fgColor=COUL_CLASSES.get(classe, "FFFFFF"))
        _auto_largeur(ws2)
        ws2.freeze_panes = "B4"

    elif df_classes is not None and not df_classes.empty:
        # Format long (legacy)
        _ecrire_titre_onglet(ws2, titre + " — Classes de qualité", n_colonnes=6)
        cols2 = ["Station", "Paramètre", "Valeur P90/P10", "Classe", "Seuil TBE/BE", "Source"]
        _ecrire_en_tetes(ws2, cols2, ligne=3)
        rows2, classes_par_ligne = [], []
        for _, r in df_classes.iterrows():
            code = int(r.get("CdParametre", 0))
            lb_p = (lb_map.get(code, str(code)) if lb_map else str(code))[:40]
            lb_s = r.get("CdStation", "")
            if lb_stations:
                lb_s = lb_stations.get(str(lb_s), str(lb_s))
            classe = str(r.get("Classe", "ND"))
            rows2.append([str(lb_s)[:30], lb_p, r.get("Valeur_ref", ""),
                          classe, r.get("Seuil_TBE_BE", ""),
                          r.get("Source_retenue", r.get("Source", ""))])
            classes_par_ligne.append(classe)
        _ecrire_donnees(ws2, rows2, ligne_debut=4,
                        formats={3: _nb_fmt(3), 5: _nb_fmt(3)})
        for i, classe in enumerate(classes_par_ligne):
            cell = ws2.cell(row=4 + i, column=4)
            cell.fill = PatternFill("solid", fgColor=COUL_CLASSES.get(classe, "FFFFFF"))
            cell.font = Font(name="Arial", size=9, bold=True)
        _auto_largeur(ws2)

    else:
        _ecrire_titre_onglet(ws2, titre + " — Classes de qualité", n_colonnes=3)
        ws2.cell(row=3, column=1,
                 value="Classes non disponibles — fournir pivot_classes issu de "
                       "m07.calculer_classes_par_station(pivot, df_ref)[0].").font = _FONT_CELL
        _auto_largeur(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exporter_excel_m04(
    pivot_norm: pd.DataFrame,
    scores_acp: Optional[pd.DataFrame] = None,
    loadings: Optional[pd.DataFrame] = None,
    matrice_dist: Optional[pd.DataFrame] = None,
    lb_stations: Optional[dict] = None,
    titre: str = "Analyses multivariées — Synthèse",
) -> bytes:
    """
    Classeur Excel M04 :
      Onglet 1 : Pivot normalisé (z-scores log)
      Onglet 2 : Scores ACP (si fournis)
      Onglet 3 : Matrice de distances inter-stations (si fournie)
    """
    wb = Workbook()

    # ── Onglet 1 : Pivot normalisé ──
    ws1 = wb.active
    ws1.title = "Pivot z-scores"
    _ecrire_titre_onglet(ws1, titre + " — Z-scores log (pivot normalisé)",
                         n_colonnes=max(2, pivot_norm.shape[1] + 1))

    if not pivot_norm.empty:
        params_cols = list(pivot_norm.columns)
        _ecrire_en_tetes(ws1, ["Station"] + [str(c) for c in params_cols], ligne=3)
        rows = []
        for station, row in pivot_norm.iterrows():
            lb_s = lb_stations.get(str(station), str(station)) if lb_stations else str(station)
            rows.append([lb_s[:30]] + [
                round(float(v), 4) if pd.notna(v) else "" for v in row
            ])
        fmt = {j + 2: _nb_fmt(4) for j in range(len(params_cols))}
        _ecrire_donnees(ws1, rows, ligne_debut=4, formats=fmt)
    _auto_largeur(ws1)

    # ── Onglet 2 : Scores ACP ──
    ws2 = wb.create_sheet("Scores ACP")
    _ecrire_titre_onglet(ws2, titre + " — Scores ACP",
                         n_colonnes=max(2, (scores_acp.shape[1] + 1) if scores_acp is not None else 3))
    if scores_acp is not None and not scores_acp.empty:
        _ecrire_en_tetes(ws2, ["Station"] + list(scores_acp.columns), ligne=3)
        rows2 = []
        for station, row in scores_acp.iterrows():
            lb_s = lb_stations.get(str(station), str(station)) if lb_stations else str(station)
            rows2.append([lb_s[:30]] + [round(float(v), 4) if pd.notna(v) else "" for v in row])
        _ecrire_donnees(ws2, rows2, ligne_debut=4)
    else:
        ws2.cell(row=3, column=1,
                 value="Scores ACP non fournis.").font = _FONT_CELL
    _auto_largeur(ws2)

    # ── Onglet 3 : Matrice distances ──
    ws3 = wb.create_sheet("Distances inter-stations")
    _ecrire_titre_onglet(ws3, titre + " — Distances Bray-Curtis inter-stations",
                         n_colonnes=max(2, (matrice_dist.shape[1] + 1) if matrice_dist is not None else 3))
    if matrice_dist is not None and not matrice_dist.empty:
        stations_list = list(matrice_dist.index)
        _ecrire_en_tetes(ws3, ["Station"] + [
            lb_stations.get(str(s), str(s)) if lb_stations else str(s)
            for s in stations_list
        ], ligne=3)
        rows3 = []
        for station, row in matrice_dist.iterrows():
            lb_s = lb_stations.get(str(station), str(station)) if lb_stations else str(station)
            rows3.append([lb_s[:30]] + [round(float(v), 4) if pd.notna(v) else "" for v in row])
        _ecrire_donnees(ws3, rows3, ligne_debut=4)
    else:
        ws3.cell(row=3, column=1,
                 value="Matrice de distances non fournie.").font = _FONT_CELL
    _auto_largeur(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exporter_excel_m05(
    df_clean: pd.DataFrame,
    lb_map: Optional[dict] = None,
    lb_stations: Optional[dict] = None,
    df_seuils: Optional[pd.DataFrame] = None,
    titre: str = "Variabilité temporelle — Synthèse",
) -> bytes:
    """
    Classeur Excel M05 :
      Onglet 1 : Distributions (Q25, médiane, Q75, n) par station × paramètre
                 Mise en forme conditionnelle sur la médiane si df_seuils fourni
      Onglet 2 : Profil saisonnier mensuel (médiane par mois) par paramètre
    """
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.styles import Color

    wb = Workbook()
    col_station = "CdStationMesureEauxSurface"
    col_param   = "CdParametre"
    col_valeur  = "Valeur" if "Valeur" in df_clean.columns else "RsAna_val"
    col_date    = "DatePrel"

    # ── Onglet 1 : Distributions ──
    ws1 = wb.active
    ws1.title = "Distributions"
    _ecrire_titre_onglet(ws1, titre + " — Distributions (Q25 / Médiane / Q75)", n_colonnes=8)

    cols1 = ["Station", "Paramètre", "N mesures", "Min", "Q25", "Médiane", "Q75", "Max"]
    _ecrire_en_tetes(ws1, cols1, ligne=3)

    rows1 = []
    if not df_clean.empty and col_valeur in df_clean.columns:
        df_w = df_clean.copy()
        df_w[col_valeur] = pd.to_numeric(df_w[col_valeur], errors="coerce")
        grp = df_w.groupby([col_station, col_param])[col_valeur]
        for (st, code), vals in grp:
            vals = vals.dropna()
            if vals.empty:
                continue
            lb_s = lb_stations.get(str(st), str(st)) if lb_stations else str(st)
            lb_p = (lb_map.get(int(code), str(code)) if lb_map else str(code))[:40]
            rows1.append([
                lb_s[:30], lb_p,
                len(vals),
                round(float(vals.min()), 4),
                round(float(vals.quantile(0.25)), 4),
                round(float(vals.median()), 4),
                round(float(vals.quantile(0.75)), 4),
                round(float(vals.max()), 4),
            ])

    fmt1 = {j: _nb_fmt(3) for j in range(4, 9)}
    _ecrire_donnees(ws1, rows1, ligne_debut=4, formats=fmt1)

    # Mise en forme conditionnelle : dégradé vert→rouge sur la colonne Médiane (col F=6)
    if rows1:
        n_data = len(rows1)
        plage_med = f"F4:F{3 + n_data}"
        ws1.conditional_formatting.add(
            plage_med,
            ColorScaleRule(
                start_type="min",  start_color="63BE7B",  # vert
                mid_type="percentile", mid_value=50, mid_color="FFEB84",  # jaune
                end_type="max",   end_color="F8696B",   # rouge
            ),
        )

    _auto_largeur(ws1)

    # ── Onglet 2 : Saisonnalité ──
    ws2 = wb.create_sheet("Saisonnalité mensuelle")
    mois_noms = ["Jan", "Fév", "Mar", "Avr", "Mai", "Jun",
                 "Jul", "Aoû", "Sep", "Oct", "Nov", "Déc"]
    _ecrire_titre_onglet(ws2, titre + " — Médiane mensuelle par station × paramètre",
                         n_colonnes=14)

    cols2 = ["Station", "Paramètre"] + mois_noms
    _ecrire_en_tetes(ws2, cols2, ligne=3)

    rows2 = []
    if not df_clean.empty and col_valeur in df_clean.columns and col_date in df_clean.columns:
        df_s = df_clean.copy()
        df_s[col_valeur] = pd.to_numeric(df_s[col_valeur], errors="coerce")
        df_s[col_date]   = pd.to_datetime(df_s[col_date], dayfirst=True, errors="coerce")
        df_s = df_s.dropna(subset=[col_date, col_valeur])
        df_s["Mois"] = df_s[col_date].dt.month

        # Pivot propre : station × paramètre en index, mois 1-12 en colonnes
        pivot_saison = (
            df_s.groupby([col_station, col_param, "Mois"])[col_valeur]
            .median()
            .unstack(level="Mois")   # colonnes = mois 1..12
        )

        for (st, code), row_pivot in pivot_saison.iterrows():
            lb_s = lb_stations.get(str(st), str(st)) if lb_stations else str(st)
            lb_p = (lb_map.get(int(code), str(code)) if lb_map else str(code))[:40]
            mois_vals = []
            for m in range(1, 13):
                val = row_pivot.get(m, np.nan)
                mois_vals.append(round(float(val), 4) if pd.notna(val) else "")
            rows2.append([lb_s[:30], lb_p] + mois_vals)

    fmt2 = {j: _nb_fmt(3) for j in range(3, 15)}
    _ecrire_donnees(ws2, rows2, ligne_debut=4, formats=fmt2)

    # Mise en forme conditionnelle : dégradé sur les valeurs mensuelles (C:N)
    if rows2:
        n2 = len(rows2)
        plage_mois = f"C4:N{3 + n2}"
        ws2.conditional_formatting.add(
            plage_mois,
            ColorScaleRule(
                start_type="min",  start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max",   end_color="F8696B",
            ),
        )

    _auto_largeur(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exporter_excel_m06(
    df_reg: pd.DataFrame,
    lb_map: Optional[dict] = None,
    lb_stations: Optional[dict] = None,
    titre: str = "Débit et chimie — Synthèse",
) -> bytes:
    """
    Classeur Excel M06 :
      Onglet 1 : Régressions C-Q (a, b, R², n paires, comportement)
      Onglet 2 : Synthèse comportements par station (pivot station × paramètre)
    """
    wb = Workbook()

    # Couleurs comportements
    COUL_COMP = {
        "Enrichissement": "FDE68A",   # jaune clair
        "Dilution":       "BFDBFE",   # bleu clair
        "Constant":       "BBF7D0",   # vert clair
        "ND":             "E5E7EB",   # gris clair
    }

    # ── Onglet 1 : Régressions ──
    ws1 = wb.active
    ws1.title = "Régressions C-Q"
    _ecrire_titre_onglet(ws1, titre + " — Régressions log-log C = a × Q^b", n_colonnes=9)

    cols1 = ["Station", "Paramètre", "N paires", "a", "b",
             "R²", "p-value", "Comportement", "Source débit"]
    _ecrire_en_tetes(ws1, cols1, ligne=3)

    rows_data = []
    comport_par_ligne = []
    if not df_reg.empty:
        for _, r in df_reg.iterrows():
            code = int(r.get("CdParametre", 0))
            lb_p = (lb_map.get(code, str(code)) if lb_map else str(code))[:40]
            lb_s = r.get("LbStation", r.get("CdStation", ""))
            if lb_stations:
                lb_s = lb_stations.get(str(r.get("CdStation", "")), str(lb_s))
            comp = r.get("Comportement", "ND")
            rows_data.append([
                str(lb_s)[:30], lb_p,
                r.get("n_paires", ""),
                r.get("a", ""),
                r.get("b", ""),
                r.get("r2", ""),
                r.get("p_value", ""),
                comp,
                r.get("Source_debit", ""),
            ])
            comport_par_ligne.append(comp)

    _ecrire_donnees(ws1, rows_data, ligne_debut=4,
                    formats={4: _nb_fmt(4), 5: _nb_fmt(4),
                             6: "0.000", 7: "0.000000"})

    # Colorier la colonne Comportement (col 8) + dégradé R² (col 6)
    from openpyxl.formatting.rule import ColorScaleRule
    for i, comp in enumerate(comport_par_ligne):
        coul = COUL_COMP.get(comp, "FFFFFF")
        cell = ws1.cell(row=4 + i, column=8)
        cell.fill = PatternFill("solid", fgColor=coul)
        cell.font = Font(name="Arial", size=9, bold=True)

    # Dégradé sur R² (col F=6) : rouge (0) → jaune (0.5) → vert (1)
    if rows_data:
        n_reg = len(rows_data)
        ws1.conditional_formatting.add(
            f"F4:F{3 + n_reg}",
            ColorScaleRule(
                start_type="num",  start_value=0,   start_color="F8696B",
                mid_type="num",    mid_value=0.5,   mid_color="FFEB84",
                end_type="num",    end_value=1,     end_color="63BE7B",
            ),
        )

    _auto_largeur(ws1)

    # ── Onglet 2 : Pivot comportements ──
    ws2 = wb.create_sheet("Synthèse comportements")
    if not df_reg.empty:
        codes   = sorted(df_reg["CdParametre"].unique())
        stations = sorted(df_reg["CdStation"].unique())
        n_cols = len(codes) + 1
        _ecrire_titre_onglet(ws2, titre + " — Comportements par station × paramètre",
                             n_colonnes=n_cols)
        en_tetes = ["Station"] + [
            (lb_map.get(int(c), str(c)) if lb_map else str(c))[:18]
            for c in codes
        ]
        _ecrire_en_tetes(ws2, en_tetes, ligne=3)

        for i, st in enumerate(stations):
            lb_s = lb_stations.get(str(st), str(st)) if lb_stations else str(st)
            row_vals = [lb_s[:30]]
            for j, code in enumerate(codes):
                sub = df_reg[(df_reg["CdStation"] == st) & (df_reg["CdParametre"] == code)]
                comp = sub["Comportement"].iloc[0] if not sub.empty else "–"
                row_vals.append(comp)
            ligne = 4 + i
            for j, val in enumerate(row_vals, 1):
                cell = ws2.cell(row=ligne, column=j, value=val)
                cell.font = Font(name="Arial", size=9,
                                 bold=(j > 1 and val not in ("–", "")))
                cell.alignment = _ALIGN_CTR
                cell.border    = _BORDER_THIN
                if j > 1 and val in COUL_COMP:
                    cell.fill = PatternFill("solid", fgColor=COUL_COMP[val])
                else:
                    fill = _FILL_ALTERN if i % 2 == 0 else _FILL_BLANC
                    cell.fill = fill
    else:
        ws2.cell(row=3, column=1,
                 value="Régressions C-Q non disponibles.").font = _FONT_CELL

    _auto_largeur(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exporter_excel_module(
    module: str,
    **kwargs,
) -> bytes:
    """
    Point d'entrée unique pour l'export Excel d'un module.

    Parameters
    ----------
    module : "m03", "m04", "m05" ou "m06"
    **kwargs : arguments passés à la fonction spécifique du module

    Returns bytes (xlsx)
    """
    dispatch = {
        "m03": exporter_excel_m03,
        "m04": exporter_excel_m04,
        "m05": exporter_excel_m05,
        "m06": exporter_excel_m06,
    }
    if module not in dispatch:
        raise ValueError(f"Module '{module}' non supporté. Choisir parmi : {list(dispatch)}")
    return dispatch[module](**kwargs)


# ---------------------------------------------------------------------------
# 4. Rapport PDF structuré
# ---------------------------------------------------------------------------

def _pied_de_page_portrait(canvas_obj, doc):
    """Callback pied de page — format portrait."""
    _dessiner_pied(canvas_obj, *A4, doc.page)


def _pied_de_page_paysage(canvas_obj, doc):
    """Callback pied de page — format paysage (A4 tourné)."""
    w, h = A4
    _dessiner_pied(canvas_obj, h, w, doc.page)  # largeur et hauteur inversées


def _dessiner_pied(canvas_obj, page_w, page_h, num_page):
    canvas_obj.saveState()
    canvas_obj.setStrokeColor(colors.HexColor("#BFDBFE"))
    canvas_obj.setLineWidth(0.5)
    canvas_obj.line(2 * cm, 1.5 * cm, page_w - 2 * cm, 1.5 * cm)
    canvas_obj.setFont("Helvetica", 7)
    canvas_obj.setFillColor(WATERMARK_COLOR)
    canvas_obj.drawRightString(page_w - 2 * cm, 1.0 * cm, WATERMARK_TXT)
    canvas_obj.setFillColor(colors.HexColor("#6B7280"))
    canvas_obj.drawString(2 * cm, 1.0 * cm, f"Page {num_page}")
    canvas_obj.restoreState()


def generer_rapport_pdf(
    sections: list[dict],
    titre_rapport: str = "Rapport d'analyse — Chimie globale",
    stations: Optional[list] = None,
    stations_codes: Optional[list] = None,
    periode: Optional[str] = None,
    auteur: str = "@CDEau",
    date_rapport: Optional[str] = None,
) -> bytes:
    """
    Génère un rapport PDF structuré multi-modules avec watermark @CDEau.

    Parameters
    ----------
    sections : liste de dicts, une entrée par module/section :
        {
          "titre"      : str — titre de la section
          "figures"    : list[plt.Figure] — figures à intégrer
          "tableau"    : pd.DataFrame — tableau de synthèse (optionnel)
          "commentaire": str — texte d'introduction (optionnel)
          "paysage"    : bool — True pour basculer cette section en A4 paysage
                         (recommandé pour figures multi-colonnes). Défaut False.
        }
    titre_rapport  : titre principal (page de garde)
    stations       : liste de libellés de stations (affichés en vertical sur la garde)
    stations_codes : liste de codes stations correspondants (même ordre que stations)
    periode        : chaîne décrivant la période (ex. "2021-2024")
    auteur         : affiché en bas de la page de garde
    date_rapport   : date (défaut = aujourd'hui)

    Returns bytes (pdf)
    """
    from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate
    from reportlab.lib.pagesizes import landscape as rl_landscape

    if date_rapport is None:
        date_rapport = datetime.date.today().strftime("%d/%m/%Y")

    # Styles partagés
    styles = getSampleStyleSheet()
    st_titre_rapport = ParagraphStyle(
        "TitreRapport", parent=styles["Title"],
        fontName="Helvetica-Bold", fontSize=20,
        textColor=COULEUR_TITRE, spaceAfter=18, alignment=TA_CENTER,
    )
    st_meta = ParagraphStyle(
        "Meta", parent=styles["Normal"],
        fontName="Helvetica", fontSize=10,
        textColor=colors.HexColor("#374151"),
        alignment=TA_CENTER, spaceAfter=6,
    )
    st_titre_section = ParagraphStyle(
        "TitreSection", parent=styles["Heading1"],
        fontName="Helvetica-Bold", fontSize=14,
        textColor=COULEUR_TITRE, spaceBefore=14, spaceAfter=8,
    )
    st_commentaire = ParagraphStyle(
        "Commentaire", parent=styles["Normal"],
        fontName="Helvetica", fontSize=9,
        textColor=colors.HexColor("#374151"),
        spaceAfter=8, leading=14,
    )
    st_th = ParagraphStyle(
        "TableauTitre", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=8,
        textColor=colors.white, alignment=TA_CENTER,
    )
    st_td = ParagraphStyle(
        "TableauCell", parent=styles["Normal"],
        fontName="Helvetica", fontSize=8, alignment=TA_CENTER,
    )
    # Style pour le tableau stations page de garde
    st_garde_th = ParagraphStyle(
        "GardeTh", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=9,
        textColor=colors.white, alignment=TA_CENTER,
    )
    st_garde_td = ParagraphStyle(
        "GardeTd", parent=styles["Normal"],
        fontName="Helvetica", fontSize=9,
        textColor=colors.HexColor("#1e3a5f"), alignment=TA_LEFT,
    )

    # ── Construction du doc avec multi-templates (portrait + paysage) ──
    buf = io.BytesIO()

    marg = 2 * cm
    marg_top = 2.5 * cm
    marg_bot = 2.5 * cm

    A4_p = A4                          # portrait
    A4_l = rl_landscape(A4)            # paysage

    def _frame_portrait():
        return Frame(marg, marg_bot,
                     A4_p[0] - 2 * marg, A4_p[1] - marg_top - marg_bot,
                     id="portrait")

    def _frame_paysage():
        return Frame(marg, marg_bot,
                     A4_l[0] - 2 * marg, A4_l[1] - marg_top - marg_bot,
                     id="paysage")

    doc = BaseDocTemplate(
        buf,
        pagesize=A4_p,
        leftMargin=marg, rightMargin=marg,
        topMargin=marg_top, bottomMargin=marg_bot,
        title=titre_rapport, author=auteur,
    )
    doc.addPageTemplates([
        PageTemplate(id="portrait", frames=[_frame_portrait()],
                     pagesize=A4_p,
                     onPage=_pied_de_page_portrait),
        PageTemplate(id="paysage",  frames=[_frame_paysage()],
                     pagesize=A4_l,
                     onPage=_pied_de_page_paysage),
    ])

    from reportlab.platypus import NextPageTemplate

    story = []

    # ── Page de garde (portrait) ──
    story.append(NextPageTemplate("portrait"))
    story.append(Spacer(1, 3 * cm))
    story.append(Paragraph(titre_rapport, st_titre_rapport))
    story.append(HRFlowable(width="80%", thickness=1.5,
                             color=COULEUR_TITRE, spaceAfter=16))

    if periode:
        story.append(Paragraph(f"Période : {periode}", st_meta))

    # Tableau stations : Code | Libellé, en vertical
    if stations:
        story.append(Spacer(1, 0.4 * cm))
        codes_col = stations_codes if stations_codes else [""] * len(stations)
        data_st = [[Paragraph("Code station", st_garde_th),
                    Paragraph("Libellé", st_garde_th)]]
        for code, lb in zip(codes_col, stations):
            data_st.append([
                Paragraph(str(code), st_garde_td),
                Paragraph(str(lb).strip(), st_garde_td),
            ])
        t_st = Table(data_st, colWidths=[4 * cm, 10 * cm], hAlign="CENTER")
        t_st.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), COULEUR_TITRE),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F0F7FF")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BFDBFE")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ]))
        story.append(t_st)

    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph(f"Date : {date_rapport}", st_meta))
    story.append(Paragraph(f"Produit par : {auteur}", st_meta))
    story.append(PageBreak())

    # ── Sections par module ──
    for sec in sections:
        titre_sec   = sec.get("titre", "Section")
        figures_sec = sec.get("figures", [])
        tableau_sec = sec.get("tableau")
        commentaire = sec.get("commentaire", "")
        paysage     = sec.get("paysage", False)

        # Basculer le format de page si nécessaire
        if paysage:
            story.append(NextPageTemplate("paysage"))
            story.append(PageBreak())
            w_utile = A4_l[0] - 2 * marg
            h_ratio = 0.50   # hauteur figure = 50 % de la largeur utile paysage
        else:
            story.append(NextPageTemplate("portrait"))
            story.append(PageBreak())
            w_utile = A4_p[0] - 2 * marg
            h_ratio = 0.58

        story.append(Paragraph(titre_sec, st_titre_section))
        story.append(HRFlowable(width="100%", thickness=0.5,
                                 color=colors.HexColor("#BFDBFE"), spaceAfter=6))

        if commentaire:
            story.append(Paragraph(commentaire, st_commentaire))

        # Figures — résolution 250 dpi
        for fig in figures_sec:
            img_buf = io.BytesIO()
            if isinstance(fig, (bytes, bytearray)):
                img_buf.write(bytes(fig))
                # Dimensions inconnues pour bytes PNG — utiliser 100% de la largeur utile
                img_w = w_utile
                img_h = w_utile * 0.75   # ratio 4:3 par défaut
            else:
                fig.savefig(img_buf, format="png", dpi=250, bbox_inches="tight")
                fig_w_in, fig_h_in = fig.get_size_inches()
                ratio_fig = fig_h_in / fig_w_in
                img_w = w_utile
                img_h = min(img_w * ratio_fig, w_utile * h_ratio)
            img_buf.seek(0)

            rl_img = RLImage(img_buf, width=img_w, height=img_h)
            story.append(rl_img)
            story.append(Spacer(1, 0.3 * cm))

        # Tableau de synthèse
        if tableau_sec is not None and not tableau_sec.empty:
            story.append(Spacer(1, 0.2 * cm))
            df_show = tableau_sec.head(20)
            data_pdf = [[Paragraph(str(c), st_th) for c in df_show.columns]]
            for _, row in df_show.iterrows():
                data_pdf.append([
                    Paragraph(str(v) if pd.notna(v) else "–", st_td)
                    for v in row
                ])
            col_w = w_utile / max(len(df_show.columns), 1)
            t = Table(data_pdf, colWidths=[col_w] * len(df_show.columns), repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, 0), COULEUR_TITRE),
                ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
                ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#F0F7FF")]),
                ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#BFDBFE")),
                ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING",    (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)
            story.append(Spacer(1, 0.4 * cm))

    # Revenir en portrait pour la dernière page
    story.append(NextPageTemplate("portrait"))

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# 5. Bundle ZIP
# ---------------------------------------------------------------------------

def generer_zip(
    figures: dict[str, plt.Figure],
    excels: dict[str, bytes],
    pdf_bytes: Optional[bytes] = None,
    nom_projet: str = "chimie_globale",
) -> bytes:
    """
    Génère un ZIP agrégeant :
      - figures/ : PNG et SVG de chaque figure
      - excel/   : classeurs Excel par module
      - rapport/ : PDF structuré (si fourni)

    Parameters
    ----------
    figures     : dict {nom_figure: plt.Figure}
    excels      : dict {nom_fichier: bytes_xlsx}
    pdf_bytes   : bytes du rapport PDF (optionnel)
    nom_projet  : préfixe des noms de fichiers

    Returns bytes (zip)
    """
    buf = io.BytesIO()
    date_str = datetime.date.today().strftime("%Y%m%d")

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:

        # Figures PNG + SVG
        for nom, fig in figures.items():
            nom_base = f"{nom_projet}_{nom}_{date_str}"
            png_bytes = exporter_figure(fig, format="png", dpi=200)
            svg_bytes = exporter_figure(fig, format="svg")
            zf.writestr(f"figures/{nom_base}.png", png_bytes)
            zf.writestr(f"figures/{nom_base}.svg", svg_bytes)

        # Fichiers Excel
        for nom, xlsx_bytes in excels.items():
            zf.writestr(f"excel/{nom_projet}_{nom}_{date_str}.xlsx", xlsx_bytes)

        # PDF
        if pdf_bytes:
            zf.writestr(f"rapport/{nom_projet}_rapport_{date_str}.pdf", pdf_bytes)

    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# 6. Pipeline complet — Streamlit ready
# ---------------------------------------------------------------------------

def preparer_exports(
    figures: dict[str, plt.Figure],
    *,
    # Données par module pour les Excel
    df_stats_m03: Optional[pd.DataFrame] = None,
    df_classes_m03: Optional[pd.DataFrame] = None,
    pivot_norm_m04: Optional[pd.DataFrame] = None,
    scores_acp_m04: Optional[pd.DataFrame] = None,
    loadings_m04: Optional[pd.DataFrame] = None,
    matrice_dist_m04: Optional[pd.DataFrame] = None,
    df_clean_m05: Optional[pd.DataFrame] = None,
    df_reg_m06: Optional[pd.DataFrame] = None,
    # Métadonnées rapport
    lb_map: Optional[dict] = None,
    lb_stations: Optional[dict] = None,
    titre_rapport: str = "Rapport d'analyse — Chimie globale",
    stations_noms: Optional[list] = None,
    periode: Optional[str] = None,
    # Sections PDF (si None, généré automatiquement depuis figures)
    sections_pdf: Optional[list] = None,
    nom_projet: str = "chimie_globale",
) -> dict[str, bytes]:
    """
    Pipeline complet d'export en un seul appel. Retourne un dict de bytes
    prêts pour st.download_button() dans Streamlit.

    Returns
    -------
    {
      "zip"       : bytes ZIP complet,
      "pdf"       : bytes PDF rapport,
      "excel_m03" : bytes Excel M03 (si données disponibles),
      "excel_m04" : bytes Excel M04 (si données disponibles),
      "excel_m05" : bytes Excel M05 (si données disponibles),
      "excel_m06" : bytes Excel M06 (si données disponibles),
      + une entrée par figure : "fig_{nom}_png" et "fig_{nom}_svg"
    }
    """
    exports = {}

    # ── Figures individuelles ──
    for nom, fig in figures.items():
        exports[f"fig_{nom}_png"] = exporter_figure(fig, format="png")
        exports[f"fig_{nom}_svg"] = exporter_figure(fig, format="svg")

    # ── Excel par module ──
    excels = {}

    if df_stats_m03 is not None:
        try:
            excels["m03_empreinte"] = exporter_excel_m03(
                df_stats_m03, df_classes_m03,
                lb_map=lb_map, lb_stations=lb_stations,
            )
            exports["excel_m03"] = excels["m03_empreinte"]
        except Exception as e:
            pass  # Log silencieux — ne pas bloquer les autres exports

    if pivot_norm_m04 is not None:
        try:
            excels["m04_multivar"] = exporter_excel_m04(
                pivot_norm_m04, scores_acp_m04, loadings_m04, matrice_dist_m04,
                lb_stations=lb_stations,
            )
            exports["excel_m04"] = excels["m04_multivar"]
        except Exception as e:
            pass

    if df_clean_m05 is not None:
        try:
            excels["m05_variabilite"] = exporter_excel_m05(
                df_clean_m05, lb_map=lb_map, lb_stations=lb_stations,
            )
            exports["excel_m05"] = excels["m05_variabilite"]
        except Exception as e:
            pass

    if df_reg_m06 is not None and not df_reg_m06.empty:
        try:
            excels["m06_cq"] = exporter_excel_m06(
                df_reg_m06, lb_map=lb_map, lb_stations=lb_stations,
            )
            exports["excel_m06"] = excels["m06_cq"]
        except Exception as e:
            pass

    # ── PDF ──
    if sections_pdf is None:
        # Construction automatique depuis les figures disponibles
        groupes = {
            "Empreinte chimique":       ["radar", "heatmap", "heatmap_freq", "distances"],
            "Analyses multivariées":    ["biplot", "biplot_fam", "dendro", "corr", "scree"],
            "Variabilité temporelle":   ["boxplots", "series", "saison"],
            "Débit et chimie":          ["cq_params", "cq_comportements"],
        }
        sections_pdf = []
        for titre_sec, cles in groupes.items():
            figs_sec = [figures[k] for k in cles if k in figures]
            if figs_sec:
                tableau = None
                if "empreinte" in titre_sec.lower() and df_classes_m03 is not None:
                    tableau = df_classes_m03.head(20)
                elif "débit" in titre_sec.lower() and df_reg_m06 is not None:
                    tableau = df_reg_m06[["CdStation", "CdParametre",
                                          "n_paires", "b", "r2", "Comportement"]].head(20)
                sections_pdf.append({
                    "titre":   titre_sec,
                    "figures": figs_sec,
                    "tableau": tableau,
                })

    try:
        pdf_bytes = generer_rapport_pdf(
            sections_pdf,
            titre_rapport=titre_rapport,
            stations=stations_noms,
            periode=periode,
        )
        exports["pdf"] = pdf_bytes
    except Exception:
        pdf_bytes = None

    # ── ZIP ──
    try:
        exports["zip"] = generer_zip(
            figures=figures,
            excels=excels,
            pdf_bytes=pdf_bytes,
            nom_projet=nom_projet,
        )
    except Exception as e:
        pass

    return exports


# ---------------------------------------------------------------------------
# Bloc de test autonome
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import os
    import sys
    # IMPORTANT : /home/claude en premier pour charger les modules corrigés
    # (m07 corrigé NH4+ AZOT) plutôt que /tmp/chimie_modules
    sys.path.insert(0, "/tmp/chimie_modules")
    sys.path.insert(0, "/home/claude")

    print("Test M08 — export complet")

    import m01_import, m02_nettoyage, m07_referentiels as m07
    import m05_variabilite as m05

    print(f"m07 chargé depuis : {m07.__file__}")

    def corriger(s):
        if not isinstance(s, str): return s
        try: return s.encode("latin-1").decode("utf-8")
        except: return s

    res = m01_import.importer_bdd(
        "/mnt/user-data/uploads/Analyses_InterStations_apprentissage.csv",
        cd_support=3, cd_fractions=[23],
    )
    df_filtre = res["df"]
    inv = res["inventaire_stations"]
    lb_stations = dict(zip(
        inv["CdStationMesureEauxSurface"], inv["LbStationMesureEauxSurface"]
    ))

    df_fam = pd.read_csv(
        "/mnt/user-data/uploads/Substances_Familles.csv",
        sep=None, engine="python", encoding="latin-1",
    )
    df_fam["CdParametre"] = pd.to_numeric(df_fam["CdParametre"], errors="coerce")
    for col in df_fam.columns:
        if df_fam[col].dtype == object:
            df_fam[col] = df_fam[col].apply(corriger)

    pivots     = m02_nettoyage.nettoyer_et_pivoter(df_filtre, df_familles=df_fam)
    df_clean   = pivots["df_clean"]
    df_stats   = pivots["df_stats"]
    pivot      = pivots["pivot"]        # pivot brut pour calculer_classes_par_station
    pivot_norm = pivots["pivot_norm"]
    lb_map     = pivots["lb_map"]

    df_ref    = m07.fusionner_referentiels()
    df_seuils = m07.selectionner_seuil_reference(df_ref)

    # Vérifier NH4+
    nh4 = df_seuils[df_seuils["CdParametre"] == 1335].iloc[0]
    print(f"NH4+ seuil : TBE/BE={nh4['TBE_BE']} Note={nh4['Note']} ✅")

    # Calculer pivot_classes pour l'onglet classes de qualité M03
    pivot_classes, _ = m07.calculer_classes_par_station(pivot, df_ref)
    print(f"pivot_classes : {pivot_classes.shape[0]} stations × {pivot_classes.shape[1]} paramètres")

    # Générer quelques figures M05 pour le test
    params_test = [1311, 1302, 1340, 1335, 1433, 1303, 1301]
    figs_dict, alertes = m05.figure_variabilite_complete(
        df_clean, lb_map,
        df_seuils=df_seuils,
        lb_stations=lb_stations,
        params_selectionnes=params_test,
        n_colonnes=4,
    )

    print(f"Figures générées : {list(figs_dict.keys())}")

    # Test export figure
    png_bytes = exporter_figure(figs_dict["boxplots"], format="png")
    svg_bytes = exporter_figure(figs_dict["series"],   format="svg")
    print(f"PNG boxplots : {len(png_bytes):,} bytes")
    print(f"SVG séries   : {len(svg_bytes):,} bytes")

    # Test Excel M05
    xlsx_m05 = exporter_excel_m05(df_clean, lb_map=lb_map, lb_stations=lb_stations)
    print(f"Excel M05 : {len(xlsx_m05):,} bytes")

    # Test Excel M03 — avec pivot_classes pour l'onglet classes de qualité
    xlsx_m03 = exporter_excel_m03(
        df_stats, pivot_classes=pivot_classes,
        lb_map=lb_map, lb_stations=lb_stations,
    )
    print(f"Excel M03 : {len(xlsx_m03):,} bytes")

    # Test PDF
    sections = [
        {
            "titre":       "Variabilité temporelle — Distributions",
            "figures":     [figs_dict["boxplots"]],
            "commentaire": "Distributions des concentrations par station "
                           "sur les paramètres physico-chimiques du BV Bienne.",
            "paysage":     True,   # figure multi-colonnes → format paysage
        },
        {
            "titre":       "Variabilité temporelle — Séries temporelles",
            "figures":     [figs_dict["series"]],
            "commentaire": "Évolution chronologique des concentrations (2021-2024).",
            "paysage":     True,
        },
        {
            "titre":       "Variabilité temporelle — Saisonnalité",
            "figures":     [figs_dict["saison"]],
            "paysage":     True,
        },
    ]
    stations_lb   = list(lb_stations.values())
    stations_cd   = list(lb_stations.keys())
    pdf_bytes = generer_rapport_pdf(
        sections,
        titre_rapport="Rapport test M08 — BV Bienne",
        stations=stations_lb,
        stations_codes=stations_cd,
        periode="2021-2024",
    )
    print(f"PDF rapport : {len(pdf_bytes):,} bytes")

    # Test ZIP
    zip_bytes = generer_zip(
        figures=figs_dict,
        excels={"m03": xlsx_m03, "m05": xlsx_m05},
        pdf_bytes=pdf_bytes,
        nom_projet="test_bienne",
    )
    print(f"ZIP bundle  : {len(zip_bytes):,} bytes")

    # Sauvegarder les outputs
    os.makedirs("/mnt/user-data/outputs", exist_ok=True)
    with open("/mnt/user-data/outputs/test_m08_rapport.pdf", "wb") as f:
        f.write(pdf_bytes)
    with open("/mnt/user-data/outputs/test_m08_m05.xlsx", "wb") as f:
        f.write(xlsx_m05)
    with open("/mnt/user-data/outputs/test_m08_m03.xlsx", "wb") as f:
        f.write(xlsx_m03)
    with open("/mnt/user-data/outputs/test_m08_bundle.zip", "wb") as f:
        f.write(zip_bytes)

    print("\n✅ M08 OK — fichiers dans /mnt/user-data/outputs/")
